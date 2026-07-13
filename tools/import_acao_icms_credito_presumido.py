#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Importa a ação "EXCLUSÃO DA INCIDÊNCIA DO IRPJ E DA CSLL SOBRE OS CRÉDITOS
PRESUMIDOS DE ICMS" a partir da planilha `Atos Concessivos - SOMENTE LUCRO REAL.xlsx`
(aba 'Lucro Real - RN').

Regra (definida pelo usuário):
  - Empresas em VERDE (fill C6EFCE) = JÁ AJUIZARAM a ação  -> elegibilidade com
    ja_ajuizada=true (ajuizada_por_nos fica NULL = a definir na UI).
  - Empresas em BRANCO = ainda NÃO entraram -> elegibilidade com elegivel=true
    (entram no pool elegível da ação).
  - Empresa que não existe no CRM é criada (stub: nome+cnpj+uf+municipio+metadados);
    o loop autônomo de enriquecimento RFB completa o resto depois (cnpj basta).

Idempotente: re-rodar não duplica (ação por nome, empresa por CNPJ, elegibilidade
por (empresa_id, acao_id)). Requer as colunas de ajuizamento (migration
20260713000000) aplicadas.

Uso:
  . tools\\pje-env.local.ps1                       # seta SUPABASE_URL + SERVICE_ROLE_KEY
  python tools\\import_acao_icms_credito_presumido.py --dry     # preview, não grava
  python tools\\import_acao_icms_credito_presumido.py           # grava
"""
import os, re, sys, json, argparse
import requests
import openpyxl

XLSX = r"C:\Users\Gabriel\Desktop\FREIRETAX\Atos Concessivos - SOMENTE LUCRO REAL.xlsx"
SHEET = "Lucro Real - RN"
GREEN = "00C6EFCE"
OWNER = "e2f9e03e-e229-4542-a21c-b97f00792803"  # profile gabriel
ACAO_NOME = "EXCLUSÃO DA INCIDÊNCIA DO IRPJ E DA CSLL SOBRE OS CRÉDITOS PRESUMIDOS DE ICMS"
FONTE = "planilha_atos_concessivos_icms_lucro_real_rn"

UF_MAP = {
    "ACRE": "AC", "ALAGOAS": "AL", "AMAPA": "AP", "AMAZONAS": "AM", "BAHIA": "BA",
    "CEARA": "CE", "DISTRITO FEDERAL": "DF", "ESPIRITO SANTO": "ES", "GOIAS": "GO",
    "MARANHAO": "MA", "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS",
    "MINAS GERAIS": "MG", "PARA": "PA", "PARAIBA": "PB", "PARANA": "PR",
    "PERNAMBUCO": "PE", "PIAUI": "PI", "RIO DE JANEIRO": "RJ",
    "RIO GRANDE DO NORTE": "RN", "RIO GRANDE DO SUL": "RS", "RONDONIA": "RO",
    "RORAIMA": "RR", "SANTA CATARINA": "SC", "SAO PAULO": "SP", "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


def clean_cnpj(s):
    if not s:
        return None
    d = re.sub(r"\D", "", str(s))
    return d if len(d) == 14 else None


def mask_cnpj(d):
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def extract_name(raw):
    if not raw:
        return None
    # tira CNPJ embutido, depois pega o pedaço com nome (ignora prefixo de seção "A", "2")
    t = re.sub(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}", "", str(raw))
    parts = [p.strip() for p in t.split("\n") if len(p.strip()) > 3]
    return parts[0].strip() if parts else (t.strip() or None)


def parse_local(loc):
    """'MACAIBA - RIO GRANDE DO NORTE' -> ('MACAIBA', 'RN')"""
    if not loc:
        return None, None
    parts = [p.strip() for p in str(loc).split("-")]
    if len(parts) >= 2:
        mun = parts[0].title()
        uf = UF_MAP.get(parts[-1].strip().upper())
        return mun, uf
    return str(loc).title(), None


def read_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in range(2, ws.max_row + 1):
        emp_raw = ws.cell(row=r, column=2).value
        if not emp_raw or not str(emp_raw).strip():
            continue
        fill = ws.cell(row=r, column=2).fill
        rgb = fill.fgColor.rgb if fill and fill.fgColor else None
        cnpj = clean_cnpj(ws.cell(row=r, column=3).value)
        if not cnpj:
            m = re.search(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}", str(emp_raw))
            if m:
                cnpj = clean_cnpj(m.group(0))
        if not cnpj:
            continue
        mun, uf = parse_local(ws.cell(row=r, column=6).value)
        rows.append({
            "cnpj": cnpj,
            "nome": extract_name(emp_raw),
            "ajuizada": str(rgb) == GREEN,
            "municipio": mun,
            "uf": uf,
            "segmento": ws.cell(row=r, column=4).value,
            "cnae": ws.cell(row=r, column=5).value,
            "funcionarios": ws.cell(row=r, column=7).value,
            "faturamento": ws.cell(row=r, column=8).value,
        })
    # dedup por CNPJ (mantém ajuizada=OR)
    by_cnpj = {}
    for x in rows:
        c = x["cnpj"]
        if c in by_cnpj:
            by_cnpj[c]["ajuizada"] = by_cnpj[c]["ajuizada"] or x["ajuizada"]
        else:
            by_cnpj[c] = x
    return list(by_cnpj.values())


class Api:
    def __init__(self):
        self.url = os.environ["SUPABASE_URL"].rstrip("/")
        self.key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
        self.h = {"apikey": self.key, "Authorization": f"Bearer {self.key}",
                  "Content-Type": "application/json"}

    def get(self, path, **params):
        r = requests.get(f"{self.url}/rest/v1/{path}", headers=self.h, params=params)
        r.raise_for_status()
        return r.json()

    def get_all(self, path, **params):
        out, off = [], 0
        while True:
            h = dict(self.h)
            h["Range-Unit"] = "items"
            h["Range"] = f"{off}-{off+999}"
            r = requests.get(f"{self.url}/rest/v1/{path}", headers=h, params=params)
            r.raise_for_status()
            batch = r.json()
            out += batch
            if len(batch) < 1000:
                break
            off += 1000
        return out

    def post(self, path, body, prefer="return=representation"):
        h = dict(self.h)
        h["Prefer"] = prefer
        r = requests.post(f"{self.url}/rest/v1/{path}", headers=h, data=json.dumps(body))
        if not r.ok:
            raise RuntimeError(f"POST {path} -> {r.status_code}: {r.text[:400]}")
        return r.json() if r.text else []


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry", action="store_true", help="preview, não grava")
    args = ap.parse_args()

    if not os.environ.get("SUPABASE_URL") or not os.environ.get("SUPABASE_SERVICE_ROLE_KEY"):
        sys.exit("ERRO: rode `. tools\\pje-env.local.ps1` antes (SUPABASE_URL/SERVICE_ROLE_KEY).")

    api = Api()
    rows = read_rows()
    verdes = [x for x in rows if x["ajuizada"]]
    print(f"Planilha: {len(rows)} empresas únicas ({len(verdes)} ajuizadas, {len(rows)-len(verdes)} elegíveis)")

    # --- empresas existentes (mapa cnpj-normalizado -> id) ---
    print("Carregando empresas do CRM...")
    emps = api.get_all("empresas", select="id,cnpj")
    crm = {}
    for e in emps:
        c = clean_cnpj(e.get("cnpj"))
        if c:
            crm[c] = e["id"]
    faltantes = [x for x in rows if x["cnpj"] not in crm]
    print(f"  {len(crm)} empresas no CRM | {len(rows)-len(faltantes)} da planilha já existem | {len(faltantes)} a criar")

    # --- ação ---
    achadas = api.get("acoes_tributarias", select="id,nome", nome=f"eq.{ACAO_NOME}")
    if achadas:
        acao_id = achadas[0]["id"]
        print(f"Ação já existe: {acao_id}")
    elif args.dry:
        acao_id = "(nova — dry run)"
        print(f"Ação a CRIAR: {ACAO_NOME!r}")
    else:
        created = api.post("acoes_tributarias", {
            "nome": ACAO_NOME, "tipo": "INICIAL", "status": "Ativa",
            "user_id": OWNER, "responsavel_id": OWNER,
        })
        acao_id = created[0]["id"]
        print(f"Ação CRIADA: {acao_id}")

    if args.dry:
        print("\n[DRY RUN] Nada gravado. Faria:")
        print(f"  - criar ação (se não existe)")
        print(f"  - criar {len(faltantes)} empresas stub")
        print(f"  - upsert {len(rows)} elegibilidades ({len(verdes)} ja_ajuizada=true)")
        print("\n  amostra de empresas a criar:")
        for x in faltantes[:8]:
            print(f"    {mask_cnpj(x['cnpj'])}  {x['nome']!r}  {x['municipio']}/{x['uf']}  ajuizada={x['ajuizada']}")
        return

    # --- cria empresas faltantes ---
    criadas = 0
    for x in faltantes:
        body = {
            "user_id": OWNER, "responsavel_id": OWNER,
            "nome": x["nome"], "cnpj": mask_cnpj(x["cnpj"]),
            "uf": x["uf"], "municipio": x["municipio"],
            "metadados": {
                "fonte": FONTE, "segmento": x["segmento"], "cnae_planilha": x["cnae"],
                "faixa_funcionarios": x["funcionarios"], "faixa_faturamento": x["faturamento"],
            },
        }
        try:
            res = api.post("empresas", body)
            crm[x["cnpj"]] = res[0]["id"]
            criadas += 1
        except RuntimeError as e:
            # 409 = já existe (corrida/CNPJ formatado diferente): recupera o id
            got = api.get("empresas", select="id,cnpj", cnpj=f"eq.{mask_cnpj(x['cnpj'])}")
            if got:
                crm[x["cnpj"]] = got[0]["id"]
            else:
                print(f"  ! falha criar {x['nome']}: {e}")
    print(f"Empresas criadas: {criadas}")

    # --- upsert elegibilidade em lote ---
    payload = []
    for x in rows:
        eid = crm.get(x["cnpj"])
        if not eid:
            continue
        payload.append({
            "empresa_id": eid, "acao_id": acao_id, "user_id": OWNER,
            "elegivel": True, "status_qualificacao": "qualificada",
            "ja_ajuizada": x["ajuizada"],
            "justificativa": "Importado da planilha Atos Concessivos ICMS (Lucro Real RN).",
        })
    # upsert (merge) por (empresa_id, acao_id)
    total = 0
    for i in range(0, len(payload), 200):
        chunk = payload[i:i+200]
        api.post("elegibilidade?on_conflict=empresa_id,acao_id", chunk,
                 prefer="resolution=merge-duplicates,return=minimal")
        total += len(chunk)
    print(f"Elegibilidades upsertadas: {total} ({sum(1 for p in payload if p['ja_ajuizada'])} ajuizadas)")
    print("OK.")


if __name__ == "__main__":
    main()
